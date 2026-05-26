"""Quick smoke-test: regenerate the report from the existing DB."""
import sys
sys.path.insert(0, '.')
from src.reporting.excel_report import build_excel_report
from src.utils.paths import PROJECT_ROOT

db_path = str(PROJECT_ROOT / 'data' / 'parsed' / 'app.db')
out_path = str(PROJECT_ROOT / 'data' / 'output' / 'vendor_comparison_matrix.xlsx')

print("Building report...")
build_excel_report(out_path, db_path=db_path)
print("Done.")

import openpyxl, os
# Check summary workbook
wb = openpyxl.load_workbook(out_path)
print(f"Summary workbook sheets: {wb.sheetnames}")

# Check per-vendor files
out_dir = str(PROJECT_ROOT / 'data' / 'output')
vendor_files = [f for f in os.listdir(out_dir) if f.startswith('vendor_') and f.endswith('.xlsx')]
print(f"Vendor files: {vendor_files}")

for vf in vendor_files:
    vwb = openpyxl.load_workbook(os.path.join(out_dir, vf))
    print(f"\n  {vf} sheets: {vwb.sheetnames}")
    # Check NB01 sheet
    if 'NB01' in vwb.sheetnames:
        ws = vwb['NB01']
        print(f"  NB01 rows 3-10:")
        for r in range(3, 11):
            row_vals = [ws.cell(r, c).value for c in range(1, 7)]
            print(f"    Row {r}: {row_vals}")
