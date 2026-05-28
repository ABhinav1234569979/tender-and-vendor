from pathlib import Path
import textwrap

import fitz
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
INCOMING = ROOT / "data" / "incoming"
INCOMING.mkdir(parents=True, exist_ok=True)


SPECS = [
    {
        "Spec_ID": "SPEC-001",
        "Category": "Environment",
        "Priority": "Critical",
        "Parameter_Name": "Operating Temperature",
        "company_Requirement": "The supplied system shall operate continuously from 0 C to 55 C without thermal shutdown or performance degradation.",
        "Expected_Evidence": "Datasheet or test declaration confirming 0 C to 55 C continuous operation.",
    },
    {
        "Spec_ID": "SPEC-002",
        "Category": "Environment",
        "Priority": "High",
        "Parameter_Name": "Storage Temperature",
        "company_Requirement": "Equipment shall support safe storage from -20 C to 70 C in non-operating condition.",
        "Expected_Evidence": "Storage temperature range stated in the vendor technical sheet.",
    },
    {
        "Spec_ID": "SPEC-003",
        "Category": "Mechanical",
        "Priority": "Critical",
        "Parameter_Name": "Hydrostatic Pressure",
        "company_Requirement": "The enclosure or shell shall withstand 60 bar hydrostatic pressure during qualification testing.",
        "Expected_Evidence": "Pressure test certificate, rating statement, or type test report.",
    },
    {
        "Spec_ID": "SPEC-004",
        "Category": "Mechanical",
        "Priority": "High",
        "Parameter_Name": "Ingress Protection",
        "company_Requirement": "Outdoor equipment shall meet IP65 ingress protection or better for dust and water exposure.",
        "Expected_Evidence": "IP rating certificate or product compliance declaration.",
    },
    {
        "Spec_ID": "SPEC-005",
        "Category": "Power",
        "Priority": "High",
        "Parameter_Name": "Input Voltage Range",
        "company_Requirement": "The unit shall operate on 230 VAC input with tolerance from 180 VAC to 260 VAC.",
        "Expected_Evidence": "Power supply datasheet or electrical compliance sheet.",
    },
    {
        "Spec_ID": "SPEC-006",
        "Category": "Power",
        "Priority": "Medium",
        "Parameter_Name": "Backup Power",
        "company_Requirement": "The system shall support at least 30 minutes of backup operation using UPS or battery support.",
        "Expected_Evidence": "UPS sizing note, battery specification, or runtime calculation.",
    },
    {
        "Spec_ID": "SPEC-007",
        "Category": "Security",
        "Priority": "Critical",
        "Parameter_Name": "Authentication",
        "company_Requirement": "Administrative access shall support OAuth2 or JWT based authentication with role based access control.",
        "Expected_Evidence": "Security architecture note showing OAuth2, JWT, or equivalent RBAC mechanism.",
    },
    {
        "Spec_ID": "SPEC-008",
        "Category": "Security",
        "Priority": "Critical",
        "Parameter_Name": "Encryption In Transit",
        "company_Requirement": "All web and API traffic shall be encrypted using TLS 1.2 or higher.",
        "Expected_Evidence": "Security specification confirming TLS 1.2 or TLS 1.3.",
    },
    {
        "Spec_ID": "SPEC-009",
        "Category": "Security",
        "Priority": "High",
        "Parameter_Name": "Encryption At Rest",
        "company_Requirement": "Sensitive records and uploaded documents shall be encrypted at rest using AES-256 or an equivalent approved algorithm.",
        "Expected_Evidence": "Data protection statement, storage architecture, or encryption policy.",
    },
    {
        "Spec_ID": "SPEC-010",
        "Category": "Security",
        "Priority": "High",
        "Parameter_Name": "Audit Trail",
        "company_Requirement": "The platform shall maintain tamper evident audit logs for user actions, file uploads, pipeline runs, and report downloads.",
        "Expected_Evidence": "Audit log design, sample audit report, or platform capability statement.",
    },
    {
        "Spec_ID": "SPEC-011",
        "Category": "Security",
        "Priority": "Medium",
        "Parameter_Name": "Session Timeout",
        "company_Requirement": "Inactive user sessions shall automatically expire after 15 minutes or less.",
        "Expected_Evidence": "Application security configuration or session management description.",
    },
    {
        "Spec_ID": "SPEC-012",
        "Category": "Data Management",
        "Priority": "Critical",
        "Parameter_Name": "Data Retention",
        "company_Requirement": "Transaction logs and compliance evidence shall be retained for 180 days minimum.",
        "Expected_Evidence": "Retention policy, database lifecycle policy, or archive configuration.",
    },
    {
        "Spec_ID": "SPEC-013",
        "Category": "Data Management",
        "Priority": "High",
        "Parameter_Name": "Backup Frequency",
        "company_Requirement": "System data shall be backed up at least once every 24 hours.",
        "Expected_Evidence": "Backup schedule, backup policy, or operations runbook.",
    },
    {
        "Spec_ID": "SPEC-014",
        "Category": "Data Management",
        "Priority": "High",
        "Parameter_Name": "Recovery Point Objective",
        "company_Requirement": "The proposed system shall support an RPO of 24 hours or better.",
        "Expected_Evidence": "Disaster recovery plan or SLA statement.",
    },
    {
        "Spec_ID": "SPEC-015",
        "Category": "Data Management",
        "Priority": "High",
        "Parameter_Name": "Recovery Time Objective",
        "company_Requirement": "The proposed system shall support an RTO of 4 hours or better for critical services.",
        "Expected_Evidence": "Disaster recovery plan or service restoration commitment.",
    },
    {
        "Spec_ID": "SPEC-016",
        "Category": "Performance",
        "Priority": "Critical",
        "Parameter_Name": "Concurrent Users",
        "company_Requirement": "The application shall support at least 500 concurrent users under normal operating load.",
        "Expected_Evidence": "Performance test report, sizing note, or load test declaration.",
    },
    {
        "Spec_ID": "SPEC-017",
        "Category": "Performance",
        "Priority": "High",
        "Parameter_Name": "Response Time",
        "company_Requirement": "Dashboard screens shall return standard results within 3 seconds for normal queries.",
        "Expected_Evidence": "Performance test result or response-time SLA.",
    },
    {
        "Spec_ID": "SPEC-018",
        "Category": "Performance",
        "Priority": "High",
        "Parameter_Name": "File Processing Capacity",
        "company_Requirement": "The system shall process vendor PDF documents up to 100 MB per file.",
        "Expected_Evidence": "Product limit statement or tested file-size capacity.",
    },
    {
        "Spec_ID": "SPEC-019",
        "Category": "Integration",
        "Priority": "High",
        "Parameter_Name": "REST API",
        "company_Requirement": "The solution shall expose REST API endpoints for upload, pipeline execution, status, and report download.",
        "Expected_Evidence": "API documentation or endpoint catalogue.",
    },
    {
        "Spec_ID": "SPEC-020",
        "Category": "Integration",
        "Priority": "Medium",
        "Parameter_Name": "Webhook Support",
        "company_Requirement": "The platform should support webhook notification when a pipeline run completes.",
        "Expected_Evidence": "Integration guide or event notification capability.",
    },
    {
        "Spec_ID": "SPEC-021",
        "Category": "Reporting",
        "Priority": "Critical",
        "Parameter_Name": "Excel Comparison Matrix",
        "company_Requirement": "The system shall generate an Excel comparison matrix showing each specification against each vendor.",
        "Expected_Evidence": "Sample generated workbook or reporting module description.",
    },
    {
        "Spec_ID": "SPEC-022",
        "Category": "Reporting",
        "Priority": "High",
        "Parameter_Name": "Evidence Citation",
        "company_Requirement": "Each automated compliance decision shall include a citation or evidence excerpt from the vendor document.",
        "Expected_Evidence": "Sample output showing cited text, page number, or document reference.",
    },
    {
        "Spec_ID": "SPEC-023",
        "Category": "Reporting",
        "Priority": "High",
        "Parameter_Name": "Reviewer Override",
        "company_Requirement": "The reviewer shall be able to override an automated compliance decision with justification.",
        "Expected_Evidence": "Review workflow screenshot or functional description.",
    },
    {
        "Spec_ID": "SPEC-024",
        "Category": "Usability",
        "Priority": "Medium",
        "Parameter_Name": "Role Dashboard",
        "company_Requirement": "The user interface shall provide a dashboard for upload status, pipeline progress, results, and downloads.",
        "Expected_Evidence": "UI mockup, screenshot, or user guide.",
    },
    {
        "Spec_ID": "SPEC-025",
        "Category": "Usability",
        "Priority": "Medium",
        "Parameter_Name": "Document Preview",
        "company_Requirement": "The user interface should allow users to open or preview uploaded vendor PDF documents.",
        "Expected_Evidence": "UI screenshot or feature note.",
    },
    {
        "Spec_ID": "SPEC-026",
        "Category": "Operations",
        "Priority": "High",
        "Parameter_Name": "Deployment Mode",
        "company_Requirement": "The solution shall support local or private network deployment without mandatory public cloud processing.",
        "Expected_Evidence": "Deployment architecture or offline deployment statement.",
    },
    {
        "Spec_ID": "SPEC-027",
        "Category": "Operations",
        "Priority": "High",
        "Parameter_Name": "Logging",
        "company_Requirement": "Application logs shall include pipeline start, upload, evaluation, report generation, and download events.",
        "Expected_Evidence": "Logging design, sample log, or operational runbook.",
    },
    {
        "Spec_ID": "SPEC-028",
        "Category": "Operations",
        "Priority": "Medium",
        "Parameter_Name": "Monitoring",
        "company_Requirement": "The system should expose health or status information for backend service availability.",
        "Expected_Evidence": "Health endpoint description, monitoring note, or status API.",
    },
    {
        "Spec_ID": "SPEC-029",
        "Category": "Support",
        "Priority": "High",
        "Parameter_Name": "Support Response",
        "company_Requirement": "The vendor shall provide initial support response within 4 business hours for severity one incidents.",
        "Expected_Evidence": "Support SLA or escalation matrix.",
    },
    {
        "Spec_ID": "SPEC-030",
        "Category": "Support",
        "Priority": "Medium",
        "Parameter_Name": "Documentation",
        "company_Requirement": "The vendor shall provide administrator documentation, installation guide, and user guide.",
        "Expected_Evidence": "Document list, sample manual, or project deliverables schedule.",
    },
]


VENDOR_PROFILES = [
    {
        "file": "vendor_alpha_compliant.pdf",
        "name": "Vendor Alpha Systems",
        "positioning": "Enterprise-grade compliant proposal with strong security and operations coverage.",
        "yes": set(s["Spec_ID"] for s in SPECS),
        "partial": set(),
        "no": set(),
        "omit": set(),
    },
    {
        "file": "vendor_beta_budget.pdf",
        "name": "Vendor Beta Solutions",
        "positioning": "Budget proposal with acceptable reporting and deployment, but weak security controls.",
        "yes": {
            "SPEC-001", "SPEC-003", "SPEC-004", "SPEC-005", "SPEC-013", "SPEC-014",
            "SPEC-019", "SPEC-021", "SPEC-024", "SPEC-026", "SPEC-028", "SPEC-030",
        },
        "partial": {"SPEC-006", "SPEC-012", "SPEC-017", "SPEC-018", "SPEC-022", "SPEC-029"},
        "no": {"SPEC-007", "SPEC-008", "SPEC-009", "SPEC-010", "SPEC-011", "SPEC-015", "SPEC-016", "SPEC-020", "SPEC-023", "SPEC-025", "SPEC-027"},
        "omit": set(),
    },
    {
        "file": "vendor_gamma_secure.pdf",
        "name": "Vendor Gamma SecureWorks",
        "positioning": "Security-heavy proposal with strong audit controls but limited mechanical/environmental evidence.",
        "yes": {
            "SPEC-007", "SPEC-008", "SPEC-009", "SPEC-010", "SPEC-011", "SPEC-012",
            "SPEC-013", "SPEC-014", "SPEC-015", "SPEC-019", "SPEC-021", "SPEC-022",
            "SPEC-023", "SPEC-026", "SPEC-027", "SPEC-028", "SPEC-030",
        },
        "partial": {"SPEC-016", "SPEC-017", "SPEC-018", "SPEC-020", "SPEC-024", "SPEC-025", "SPEC-029"},
        "no": {"SPEC-001", "SPEC-002", "SPEC-003", "SPEC-004", "SPEC-005", "SPEC-006"},
        "omit": set(),
    },
    {
        "file": "vendor_delta_vague.pdf",
        "name": "Vendor Delta Innovations",
        "positioning": "Marketing-style proposal with vague language and very little hard evidence.",
        "yes": {"SPEC-021", "SPEC-024", "SPEC-028"},
        "partial": {"SPEC-019", "SPEC-022", "SPEC-025", "SPEC-030"},
        "no": {"SPEC-007", "SPEC-008", "SPEC-009", "SPEC-010", "SPEC-012", "SPEC-013", "SPEC-014", "SPEC-015", "SPEC-016", "SPEC-017", "SPEC-018", "SPEC-020", "SPEC-023", "SPEC-026", "SPEC-027", "SPEC-029"},
        "omit": {"SPEC-001", "SPEC-002", "SPEC-003", "SPEC-004", "SPEC-005", "SPEC-006", "SPEC-011"},
    },
    {
        "file": "vendor_epsilon_contradictory.pdf",
        "name": "Vendor Epsilon Controls",
        "positioning": "Technically detailed proposal with several explicit deviations and contradictions.",
        "yes": {
            "SPEC-001", "SPEC-004", "SPEC-005", "SPEC-008", "SPEC-010", "SPEC-013",
            "SPEC-019", "SPEC-021", "SPEC-022", "SPEC-024", "SPEC-026", "SPEC-027", "SPEC-028", "SPEC-030",
        },
        "partial": {"SPEC-002", "SPEC-006", "SPEC-009", "SPEC-016", "SPEC-017", "SPEC-020", "SPEC-023", "SPEC-025", "SPEC-029"},
        "no": {"SPEC-003", "SPEC-007", "SPEC-011", "SPEC-012", "SPEC-014", "SPEC-015", "SPEC-018"},
        "omit": set(),
    },
    {
        "file": "vendor_zeta_non_compliant.pdf",
        "name": "Vendor Zeta Legacy",
        "positioning": "Legacy proposal with minimal automation and several missing modern security capabilities.",
        "yes": {"SPEC-004", "SPEC-005", "SPEC-021", "SPEC-030"},
        "partial": {"SPEC-001", "SPEC-003", "SPEC-013", "SPEC-019", "SPEC-024", "SPEC-028", "SPEC-029"},
        "no": {
            "SPEC-002", "SPEC-006", "SPEC-007", "SPEC-008", "SPEC-009", "SPEC-010",
            "SPEC-011", "SPEC-012", "SPEC-014", "SPEC-015", "SPEC-016", "SPEC-017",
            "SPEC-018", "SPEC-020", "SPEC-022", "SPEC-023", "SPEC-025", "SPEC-026", "SPEC-027",
        },
        "omit": set(),
    },
]


def clean_incoming() -> None:
    """Remove generated sample input files so each run starts clean."""
    for path in INCOMING.glob("*"):
        if path.suffix.lower() in {".xlsx", ".pdf"}:
            path.unlink()


def write_master_workbook() -> Path:
    master = INCOMING / "master_spec.xlsx"
    df = pd.DataFrame(SPECS)
    # Keep the legacy parser-friendly columns first.
    df = df[
        [
            "Spec_ID",
            "Parameter_Name",
            "company_Requirement",
            "Category",
            "Priority",
            "Expected_Evidence",
        ]
    ]
    with pd.ExcelWriter(master, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Technical Checklist", index=False)

    wb = load_workbook(master)
    ws = wb["Technical Checklist"]
    ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="17324D")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = {
        "A": 14,
        "B": 28,
        "C": 82,
        "D": 20,
        "E": 14,
        "F": 70,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(master)
    return master


def yes_statement(vendor_name: str, spec: dict) -> str:
    return (
        f"{spec['Spec_ID']} COMPLIES. {vendor_name} meets the tender requirement for "
        f"{spec['Parameter_Name']}. Requirement satisfied: {spec['company_Requirement']} "
        f"Evidence offered: {spec['Expected_Evidence']}"
    )


def partial_statement(vendor_name: str, spec: dict) -> str:
    return (
        f"{spec['Spec_ID']} PARTIAL RESPONSE. {vendor_name} provides partial coverage for "
        f"{spec['Parameter_Name']}. The capability is available in principle, but final sizing, "
        f"configuration, or project-specific validation is required before full acceptance."
    )


def no_statement(vendor_name: str, spec: dict) -> str:
    return (
        f"{spec['Spec_ID']} DEVIATION. {vendor_name} does not include a confirmed response for "
        f"{spec['Parameter_Name']} in the base offer. This item is not provided as a committed "
        f"compliance item and should be treated as a gap unless clarified by the vendor."
    )


def supporting_noise(vendor_name: str) -> list[str]:
    return [
        f"{vendor_name} submits this synthetic technical proposal for evaluation against the buyer checklist.",
        "All names, values, and clauses in this document are generated sample data for development and demonstration only.",
        "The proposal includes commercial assumptions, solution architecture notes, operational responsibilities, and compliance statements.",
    ]


def write_pdf(path: Path, title: str, paragraphs: list[str]) -> None:
    doc = fitz.open()
    page = None
    y = 54
    margin_x = 50
    page_width = 545

    def new_page():
        nonlocal page, y
        page = doc.new_page()
        y = 54
        page.insert_text((margin_x, 32), title, fontsize=14)
        page.insert_text((margin_x, 790), "Synthetic sample proposal - not customer or vendor data", fontsize=8)

    new_page()
    for paragraph in paragraphs:
        wrapped = textwrap.fill(paragraph, width=95)
        rect_height = max(58, 14 * (wrapped.count("\n") + 2))
        if y + rect_height > 760:
            new_page()
        rect = fitz.Rect(margin_x, y, page_width, y + rect_height)
        page.insert_textbox(rect, wrapped, fontsize=10, lineheight=1.25)
        y += rect_height + 8

    doc.save(str(path))


def write_vendor_pdf(profile: dict) -> Path:
    specs_by_id = {s["Spec_ID"]: s for s in SPECS}
    path = INCOMING / profile["file"]
    vendor_name = profile["name"]

    paragraphs = [
        f"{vendor_name} - Technical Proposal",
        profile["positioning"],
        *supporting_noise(vendor_name),
        "Compliance response summary: COMPLIES means committed support; PARTIAL RESPONSE means conditional or configuration-dependent support; DEVIATION means not included or not confirmed.",
    ]

    for spec in SPECS:
        spec_id = spec["Spec_ID"]
        if spec_id in profile["omit"]:
            continue
        if spec_id in profile["yes"]:
            paragraphs.append(yes_statement(vendor_name, spec))
        elif spec_id in profile["partial"]:
            paragraphs.append(partial_statement(vendor_name, spec))
        elif spec_id in profile["no"]:
            paragraphs.append(no_statement(vendor_name, spec))
        else:
            paragraphs.append(partial_statement(vendor_name, spec))

    paragraphs.extend(
        [
            "The vendor confirms that final acceptance shall depend on contract terms, implementation workshop findings, and customer approval.",
            "This generated document intentionally mixes strong, weak, missing, and contradictory evidence so the parser, evaluator, report, and frontend can be tested safely.",
        ]
    )

    write_pdf(path, f"{vendor_name} Proposal", paragraphs)
    return path


def main() -> None:
    clean_incoming()
    master = write_master_workbook()
    vendor_paths = [write_vendor_pdf(profile) for profile in VENDOR_PROFILES]

    print(f"Created rich synthetic master spec at: {master}")
    print(f"Created {len(vendor_paths)} synthetic vendor PDFs:")
    for path in vendor_paths:
        print(f" - {path}")


if __name__ == "__main__":
    main()
