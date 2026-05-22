from typing import List
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font


def build_excel_report(output_path: str, db_path: str = "data/parsed/app.db") -> None:
    """Build a styled Excel report from the compliance_matrix table."""
    conn = sqlite3.connect(db_path)
    df = pd.read_sql_query("SELECT * FROM compliance_matrix", conn)
    conn.close()

    if df.empty:
        # create empty workbook
        df = pd.DataFrame(columns=["spec_id", "vendor_id", "status", "citation", "reasoning", "confidence"])

    # pivot to matrix: rows=spec_id, cols=vendor_id
    pivot = df.pivot(index="spec_id", columns="vendor_id", values="status")
    # write to excel
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pivot.to_excel(writer, sheet_name="Matrix")
        df.to_excel(writer, sheet_name="Raw", index=False)

    # apply simple styling
    wb = load_workbook(output_path)
    ws = wb["Matrix"]
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            val = (cell.value or "").upper()
            cell.alignment = Alignment(wrap_text=True)
            if val.startswith("YES"):
                cell.fill = green
            elif val.startswith("NEARLY"):
                cell.fill = yellow
            elif val.startswith("NO"):
                cell.fill = red

    wb.save(output_path)
